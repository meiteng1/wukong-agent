from langchain.tools import tool
import os
from openpyxl import load_workbook

@tool
def read_filtered_excel_tables(file_path: str = None):
    """
    强制从 .env 的 REFER_FILE_OUT_PATH 读取 Excel，
    自动生成两个筛选表（表3.1 和 表3.2）：

    返回：
    {
        "table31": [
            "桥墩,构件,部位,缺陷类型,现场照片",
            ...
        ],
        "table32": [ ... ]
    }
    """

    # --- 1. 永远使用 .env 的 REFER_FILE_OUT_PATH ---
    env_path = os.getenv("REFER_FILE_OUT_PATH")
    if not env_path:
        raise ValueError("ERROR: .env 中未设置 REFER_FILE_OUT_PATH，请检查 .env 文件。")

    excel_path = env_path  # 强制覆盖参数

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"ERROR: REFER_FILE_OUT_PATH 指向的文件不存在：{excel_path}")

    # --- 2. 使用 openpyxl 读取 XLSX ---
    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception as e:
        raise ValueError(f"无法读取 Excel 文件，请检查文件格式是否为 .xlsx。\n错误信息: {str(e)}")

    table31_lines = []
    table32_lines = []

    # --- 3. 遍历每个 sheet ---
    for ws in wb.worksheets:
        # 读取表头
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(h).strip() if h is not None else "" for h in rows[0]]

        # 列名兼容与映射
        def normalize_col(name):
            m = {
                "桥墩编号": "桥墩",
                "缺陷部位（里程/侧别）": "部位",
            }
            return m.get(name, name)

        header = [normalize_col(h) for h in header]

        # 构建列索引字典
        col_idx = {name: i for i, name in enumerate(header)}

        # 保证必需列键存在（缺失则索引为 None）
        required_cols = ["桥墩", "构件", "部位", "缺陷类型", "现场照片"]
        idx = {c: col_idx.get(c, None) for c in required_cols}

        # 逐行处理数据
        for r in rows[1:]:
            def val(col):
                i = idx[col]
                return (str(r[i]).strip() if (i is not None and r[i] is not None) else "")

            pier = val("桥墩")
            comp = val("构件")
            pos = val("部位")
            dtype = val("缺陷类型")
            photo = val("现场照片")

            line = ",".join([pier, comp, pos, dtype, photo])

            # 仅保留桥墩编号以 "HC" 开头的条目
            is_hc = str(pier).strip().startswith("HC")
            # --- 5. 筛选表 3.1：包含 #梁、#墩 ---
            if is_hc and (("#梁" in comp) or ("#墩" in comp)):
                table31_lines.append(line)
            # --- 6. 筛选表 3.2：支座系统 ---
            if is_hc and (("#防落梁块" in comp) or ("#垫石" in comp) or ("#支座板" in comp) or ("#支座" in comp)):
                table32_lines.append(line)

    # --- 8. 返回结果 ---
    return {
        "table31": table31_lines,
        "table32": table32_lines
    }
